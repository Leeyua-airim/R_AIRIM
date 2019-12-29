import feedparser
from urllib.parse import quote
from RPA_email_set import send_mail
from openpyxl import Workbook
import ssl

base_rss_url = "https://news.google.com/news/rss/headlines/section/topic"
business_rss = base_rss_url+quote('/BUSINESS.ko_kr/경제?ned=kr&hl=ko')
tech_rss = base_rss_url+quote('/SCITECH.ko_kr/과학기술?ned=kr&hl=ko')
health_rss = base_rss_url+quote('/HEALTH.ko_kr/건강?ned=kr&hl=ko')

xlsx=Workbook()
business_sheet = xlsx.create_sheet('경제')
business_sheet.append(['기사제목','링크','날짜'])
xlsx.save('test22.xlsx')

ssl._create_default_https_context = ssl._create_unverified_context
news_list=feedparser.parse(business_rss)
for news in news_list['items']:
    business_sheet.append([news['title'], news['link'],news['published']])

tech_sheet = xlsx.create_sheet('과학기술')
tech_sheet.append(['기사제목','링크','날짜'])
news_list = feedparser.parse(tech_rss)
for news in news_list['items']:
    tech_sheet.append([news['title'],news['link'], news['published']])

health_sheet = xlsx.create_sheet('건강')
health_sheet.append(['기사제목','링크','날짜'])
news_list = feedparser.parse(health_rss)
for news in news_list['items']:
    tech_sheet.append([news['title'],news['link'], news['published']])

file_name = 'E:\\hello-git-sourcetree\\R_GO\\Python_RPA\\news_list.xlsx'
xlsx.save(file_name)

#send_mail('이재화', 'brink0@naver.com','뉴스 수집 결과 입니다.',file_name)
