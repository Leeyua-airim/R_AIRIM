from os import listdir
from openpyxl import load_workbook, Workbook

#절대경로 지정
files = listdir("E:\\hello-git-sourcetree\\R_GO\\Python_RPA\\work_xlsx_data\\")
print(files)

#빈 클래스함수 생성
result_xlsx = Workbook()
#시트 활성화
result_sheet = result_xlsx.active
#파일 읽기
for myfile in files:
    #엑셀파일인 경우 제외
    if myfile[-4:] != 'xlsx':
        continue
    #읽기전용으로 파일읽어나가기
    tg_xlsx = load_workbook(myfile, read_only=True)
    tg_sheet = tg_xlsx.active
    #행 데이터 읽어들이기
    for row in tg_sheet.iter_rows():
        row_data = []
        #행을 row_data에 넣기
        for cell in row:
            row_data.append(cell.value)
        #최종적으로 result_sheet에 추가
        result_sheet.append(row_data)

result_xlsx.save('E:\\hello-git-sourcetree\\R_GO\\Python_RPA\\work_xlsx_data\\result.xlsx')
