from openpyxl import load_workbook

xlsx_file = "E:\\hello-git-sourcetree\\R_GO\\Python_RPA\\"
xlsx = load_workbook(xlsx_file+"result.xlsx", read_only =True)
sheet=xlsx.active

print(sheet['A25'].value)
print(sheet['B1'].value)

row = sheet['1']
for data in row:
    print(data.value)

xlsx=load_workbook(xlsx_file+"result.xlsx")
sheet=xlsx.active

col = sheet['A']
'''
for data in col:
    print(data.value)
'''
print('-'*10,'multi row_data call','-'*10)

rows = sheet['1:2']
for row in rows:
    for rowdata in row:
        print(rowdata.value)

print('시트 일부의 셀 데이터 읽기')
rows = sheet['A3:B5']
for row in rows:
    for cel in row:
        print(cel.value)

from openpyxl import Workbook
xlsx=Workbook()
sheet=xlsx.active

sheet['A1'] = 'my input data'
#xlsx.save('other.xlsx')

sheet.append(['A1-data','B1-data','C1-data'])
sheet.append(['A2-data','B2-data','C2-data'])

xlsx.save('other2.xlsx')

sheet = xlsx.create_sheet('new sheet')
sheet['A2'] = 'AIRIM'

xlsx.save('new_xlsx.xlsx')
