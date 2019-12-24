from RPA_my_email import send_mail
from openpyxl import load_workbook

xlsx = load_workbook("E:\\hello-git-sourcetree\\R_GO\\Python_RPA\\new_file.xlsx", read_only =True)
sheet=xlsx.active

for row in sheet.iter_rows():
    name = row[0].value
    mail = row[1].value
    status = row[3].value

    if status == '결제완료':
        contents = '결제완료가 확인되어 상세 커리큘럼을 안내해 드립니다.'
        send_mail(name,mail,contents,'E:\\hello-git-sourcetree\\R_GO\\Python_RPA\\커리큘럼.xlsx')
